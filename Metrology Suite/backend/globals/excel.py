import openpyxl
from datetime import date


class Excel:
    fname = ''

    def __init__(self, open_path, uut, template=None):
        self.date_format = f'_{date.today().strftime("%y%m%d")}'
        self.fname = f'{uut}{self.date_format}.xlsx'

        if template is None:
            self.wb = openpyxl.load_workbook(fr'{open_path}\{self.fname}')
        else:
            self.wb = openpyxl.load_workbook(fr'{open_path}\{template}.xlsx')
        self.uut = uut
        self.datasheet = self.wb.active

    def write(self, cell, info, sheet=None):
        if sheet is None:
            self.datasheet[cell] = info
        else:
            sheet[cell] = info

    def read(self, cell, sheet=None):
        if sheet is None:
            return self.datasheet[cell].value
        else:
            return sheet[cell].value

    def save(self, save_path):
        save_name = fr'{save_path}\{self.fname}'
        self.wb.save(save_name)

    def new_sheet(self, copied):
        self.wb.copy_worksheet(copied)

    def get_sheet_names(self) -> list:
        return self.wb.sheetnames

    def set_sheet_name(self, sheet, name):
        sheet.title = name
